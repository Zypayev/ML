import openpyxl
from openpyxl import load_workbook

def copy_first_row(source_file, destination_file):
    # Load the source Excel file
    source_workbook = openpyxl.load_workbook(source_file)
    source_sheet = source_workbook.active
    
    # Load the destination Excel file
    dest_workbook = openpyxl.load_workbook(destination_file)
    dest_sheet = dest_workbook.active
    
    # Clear all cells in the destination sheet
    dest_sheet.delete_rows(1, dest_sheet.max_row)
    
    # Copy the values from the first row of the source to the first row of the destination
    for col_num, cell in enumerate(source_sheet[1], start=1):
        dest_sheet.cell(row=1, column=col_num).value = cell.value
    
    # Save the changes to the destination Excel file
    dest_workbook.save(destination_file)


# Example usage:
copy_first_row("test_dataset.xlsx", "readyxl.xlsx")

# Load the user-uploaded Excel file
from openpyxl import load_workbook

# Load the "readyxl" Excel file
readyxl_workbook = load_workbook('readyxl.xlsx')
readyxl_worksheet = readyxl_workbook.active

# Define the headers in the "readyxl" file
readyxl_headers = [
    'team', 'targeted_productivity', 'smv', 'wip', 'over_time', 'incentive', 'idle_time', 'idle_men',
    'no_of_style_change', 'no_of_workers', 'month', 'quarter_Quarter1', 'quarter_Quarter2', 'quarter_Quarter3',
    'quarter_Quarter4', 'quarter_Quarter5', 'department_finishing', 'department_finishing', 'department_sweing',
    'day_Monday', 'day_Saturday', 'day_Sunday', 'day_Thursday', 'day_Tuesday', 'day_Wednesday'
]

# Load the user-uploaded Excel file
user_uploaded_workbook = load_workbook('user_uploaded.xlsx')
user_uploaded_worksheet = user_uploaded_workbook.active

# Map the headers in the user-uploaded file to the corresponding columns in the "readyxl" file
# Assume that the headers in the user-uploaded file are in the first row
uploaded_headers = [cell.value for cell in user_uploaded_worksheet[1]]
mapped_headers = [header for header in uploaded_headers if header in readyxl_headers]

# Iterate through the user-uploaded file and fill the "readyxl" file with available values
for row_idx, row in enumerate(user_uploaded_worksheet.iter_rows(min_row=2, values_only=True), start=2):
    for col_idx, value in enumerate(row, start=1):
        header = uploaded_headers[col_idx - 1]  # Adjust for 0-based index
        if header in mapped_headers:
            readyxl_col_idx = readyxl_headers.index(header) + 1  # Adjust for 1-based index
            readyxl_worksheet.cell(row=row_idx, column=readyxl_col_idx, value=value)

# Fill missing values with 0s
for row_idx, row in enumerate(readyxl_worksheet.iter_rows(min_row=2, values_only=True), start=2):
    for col_idx, value in enumerate(row, start=1):
        if value is None:
            readyxl_worksheet.cell(row=row_idx, column=col_idx, value=0)

# Save the updated "readyxl" workbook
readyxl_workbook.save('readyxl.xlsx')

print("Data from user-uploaded file has been processed and added to 'readyxl.xlsx'.")
