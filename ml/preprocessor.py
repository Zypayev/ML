from openpyxl import load_workbook

def run(filename):
    # Load the "readyxl" Excel file
    readyxl_workbook = load_workbook('./ml/readyxl.xlsx')
    readyxl_worksheet = readyxl_workbook.active  

        # Clear all cells in the destination sheet
    readyxl_worksheet.delete_rows(2, readyxl_worksheet.max_row)

    header = [
    'team', 'targeted_productivity', 'smv', 'wip', 'over_time', 'incentive', 'idle_time', 'idle_men',
    'no_of_style_change', 'no_of_workers', 'month', 'quarter_Quarter1', 'quarter_Quarter2', 'quarter_Quarter3',
    'quarter_Quarter4', 'quarter_Quarter5', 'department_finishing', 'department_finishing ', 'department_sweing',
    'day_Monday', 'day_Saturday', 'day_Sunday', 'day_Thursday', 'day_Tuesday', 'day_Wednesday'
    ]


    # Copy the values from the header list to the first row of the worksheet
    for col_num, header_value in enumerate(header, start=1):
        readyxl_worksheet.cell(row=1, column=col_num).value = header_value

    # Save the changes to the workbook
    readyxl_workbook.save('./ml/readyxl.xlsx')

    def find_header_row(sheet):
        # List of header keywords
        header_keywords = header
        
        # Search for header keywords within the first 10 rows
        for row_idx in range(1, min(11, sheet.max_row + 1)):
            row_values = [cell.value.lower() if cell.value else '' for cell in sheet[row_idx]]
            for keyword in header_keywords:
                if keyword.lower() in row_values:
                    return row_idx
        return None

    # Define the headers in the "readyxl" file
    readyxl_headers = [
        'team', 'targeted_productivity', 'smv', 'wip', 'over_time', 'incentive', 'idle_time', 'idle_men',
        'no_of_style_change', 'no_of_workers', 'month', 'quarter_Quarter1', 'quarter_Quarter2', 'quarter_Quarter3',
        'quarter_Quarter4', 'quarter_Quarter5', 'department_finishing', 'department_finishing', 'department_sweing',
        'day_Monday', 'day_Saturday', 'day_Sunday', 'day_Thursday', 'day_Tuesday', 'day_Wednesday'
    ]

    # Load the user-uploaded Excel file
    user_uploaded_workbook = load_workbook(f"./uploads/{filename}")
    user_uploaded_worksheet = user_uploaded_workbook.active

    # Map the headers in the user-uploaded file to the corresponding columns in the "readyxl" file
    # Assume that the headers in the user-uploaded file are in the first row
    header_row = find_header_row(user_uploaded_worksheet)
    uploaded_headers = [cell.value for cell in user_uploaded_worksheet[header_row]]
    mapped_headers = [header for header in uploaded_headers if header in readyxl_headers]

    # Iterate through the user-uploaded file and fill the "readyxl" file with available values
    for row_idx, row in enumerate(user_uploaded_worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=1):
        for col_idx, value in enumerate(row, start=1):
            header = uploaded_headers[col_idx - 1]  # Adjust for 0-based index
            if header in mapped_headers:
                readyxl_col_idx = readyxl_headers.index(header) + 1  # Adjust for 1-based index
                readyxl_worksheet.cell(row=row_idx + 1, column=readyxl_col_idx, value=value)

    # Fill missing values with 0s
    for row_idx, row in enumerate(readyxl_worksheet.iter_rows(min_row=2, values_only=True), start=2):
        for col_idx, value in enumerate(row, start=1):
            if value is None:
                readyxl_worksheet.cell(row=row_idx, column=col_idx, value=0)

    # Save the updated "readyxl" workbook
    readyxl_workbook.save('./ml/readyxl.xlsx')

    print("Data from user-uploaded file has been processed and added to 'readyxl.xlsx'.")
